import os
import requests
import json
from robot.libraries.BuiltIn import BuiltIn
# from robot.api import logger
from openpyxl import load_workbook

CONST_SHEET_ENV         = 'enviornment'
CONST_FMS_URL           = 2
CONST_GATEWAY           = 3
CONST_DEVICE_TYPE       = 4
CONST_APP_BUNDLE_YL     = 5
CONST_APP_BUNDLE_MC     = 6
CONST_USER              = 7
CONST_PASSWORD          = 8
CONST_SESSION_ID        = 9
CONST_APP_UPDATE        = 10
CONST_AS_SERVER_URL     = 11

CONST_SHEET_AS          = 'appserver'
CONST_AS_DEV_EUI_COL    = 1
CONST_AS_APP_EUI_COL    = 2
CONST_AS_APP_KEY_COL    = 3
CONST_AS_NODE_NAME_COL  = 4
CONST_AS_URL_COL        = 5

CONST_SHEET_DEVICE      = 'devices'

def load_xl(datafile, sheet):
    owd = os.getcwd()
    os.chdir('../Resources/testdata')
    global wb, ws
    wb = load_workbook(filename=datafile)
    ws = wb.get_sheet_by_name(sheet)
    os.chdir(owd)
    return wb, ws

######################
# SHEET: envionrment
######################
def get_url_by_header(datafile, server):
    wb, ws = load_xl(datafile, CONST_SHEET_ENV)
    
    for i in range(1, ws.max_column + 1):
        if server == ws.cell(row=1, column=i).value:
            url = ws.cell(row=CONST_FMS_URL, column=i).value
    return url

def get_gw_by_header(datafile, server):
    wb, ws = load_xl(datafile, CONST_SHEET_ENV)
    for i in range(1, ws.max_column + 1):
        if server == ws.cell(row=1, column=i).value:
            gateway = ws.cell(row=CONST_GATEWAY, column=i).value
    return gateway

def get_dev_type_by_header(datafile, server):
    wb, ws = load_xl(datafile, CONST_SHEET_ENV)
    for i in range(1, ws.max_column + 1):
        if server == ws.cell(row=1, column=i).value:
            dev_type = ws.cell(row=CONST_DEVICE_TYPE, column=i).value
    return dev_type

def get_app_bdl_yl_by_header(datafile, server):
    wb, ws = load_xl(datafile, CONST_SHEET_ENV)
    for i in range(1, ws.max_column + 1):
        if server == ws.cell(row=1, column=i).value:
            app_bundle = ws.cell(row=CONST_APP_BUNDLE_YL, column=i).value
    return app_bundle

def get_app_bdl_mc_by_header(datafile, server):
    wb, ws = load_xl(datafile, CONST_SHEET_ENV)
    for i in range(1, ws.max_column + 1):
        if server == ws.cell(row=1, column=i).value:
            app_bundle = ws.cell(row=CONST_APP_BUNDLE_MC, column=i).value
    return app_bundle

def get_user_by_header(datafile, header):
    wb, ws = load_xl(datafile, CONST_SHEET_ENV)
    for i in range(1, ws.max_column + 1):
        if header == ws.cell(row=1, column=i).value:
            user = ws.cell(row=CONST_USER, column=i).value
    return user

def get_password_by_header(datafile, header):
    wb, ws = load_xl(datafile, CONST_SHEET_ENV)
    for i in range(1, ws.max_column + 1):
        if header == ws.cell(row=1, column=i).value:
            pwd = ws.cell(row=CONST_PASSWORD, column=i).value
    return pwd

def get_session_by_header(datafile, server):
    wb, ws = load_xl(datafile, CONST_SHEET_ENV)
    for i in range(1, ws.max_column + 1):
        if server == ws.cell(row=1, column=i).value:
            sessionid = ws.cell(row=CONST_SESSION_ID, column=i).value
    return sessionid

def set_session_by_header(datafile, server):
    wb, ws = load_xl(datafile, CONST_SHEET_ENV)
    urlRequest = get_url_by_header(datafile, server)
    username = get_user_by_header(datafile, server)
    password  = get_password_by_header(datafile, server)
    sessionid = get_session_by_header(datafile, server)
    owd = os.getcwd()
    os.chdir('../Resources/testdata')
    userLoginRequest = urlRequest + '/login'
    dataValues       = {"username":username, "password":password}
    if len(sessionid) != 32:
        response         = requests.post(url=userLoginRequest,data=dataValues)
        jsonReply        = json.loads(response.text)
        retStatus        = jsonReply['response'].encode('utf-8')

        if retStatus == 'ok':
            retSessionID     = jsonReply['sessionid'].encode('utf-8')
            for i in range(1, ws.max_column + 1):
                if server == ws.cell(row=1, column=i).value:
                    ws.cell(row=CONST_SESSION_ID, column=i).value = retSessionID
                    wb.save(datafile)
                    os.chdir(owd)
                    return retSessionID
        else:
            os.chdir(owd)
            return None
    else:
        checkLoginRequest = urlRequest + '/checklogin'
        dataValues        = {"sessionid":sessionid}
        response          = requests.post(url=checkLoginRequest,data=dataValues)
        jsonReply         = json.loads(response.text)
        retStatus         = jsonReply['response'].encode('utf-8')

        if retStatus == 'ok':
            os.chdir(owd)
            return sessionid
        else:
            dataValues       = {"username":username, "password":password}
            response         = requests.post(url=userLoginRequest,data=dataValues)
            jsonReply        = json.loads(response.text)
            retStatus        = jsonReply['response'].encode('utf-8')
            if retStatus == 'ok':
                retSessionID     = jsonReply['sessionid'].encode('utf-8')
                for i in range(1, ws.max_column + 1):
                    if server == ws.cell(row=1, column=i).value:
                        ws.cell(row=CONST_SESSION_ID, column=i).value = retSessionID
                        wb.save(datafile)
                        os.chdir(owd)
                        return retSessionID
            else:
                os.chdir(owd)
                return None

def get_app_update_version(datafile, server):
    wb, ws = load_xl(datafile, CONST_SHEET_ENV)
    for i in range(1, ws.max_column + 1):
        if server == ws.cell(row=1, column=i).value:
            new_app = ws.cell(row=CONST_APP_UPDATE, column=i).value
    return new_app

def get_as_server_url(datafile, server):
    wb, ws = load_xl(datafile, CONST_SHEET_ENV)
    for i in range(1, ws.max_column + 1):
        if server == ws.cell(row=1, column=i).value:
            as_server_url = ws.cell(row=CONST_AS_SERVER_URL, column=i).value
    return as_server_url

######################
# SHEET: appserver
######################

def get_as_dev_eui(datafile):
    wb, ws = load_xl(datafile, CONST_SHEET_AS)
    as_dev_euis = []
    for i in range(2, ws.max_row +1):
        as_dev_euis.append(str(ws.cell(row=i, column =CONST_AS_DEV_EUI_COL).value))
    return as_dev_euis

def get_as_app_eui(datafile):
    wb, ws = load_xl(datafile, CONST_SHEET_AS)
    return str(ws.cell(row=2, column=CONST_AS_APP_EUI_COL).value)

def get_as_app_key(datafile):
    wb, ws = load_xl(datafile, CONST_SHEET_AS)
    return str(ws.cell(row=2, column=CONST_AS_APP_KEY_COL).value)

def get_as_node_name(datafile):
    wb, ws = load_xl(datafile, CONST_SHEET_AS)
    return str(ws.cell(row=2, column=CONST_AS_NODE_NAME_COL).value)


######################
# SHEET: devices
######################
def get_devices_by_header(datafile, dev_hdr):
    wb, ws = load_xl(datafile, CONST_SHEET_DEVICE)
    devices = []
    for i in range(1, ws.max_column + 1):
        if dev_hdr == ws.cell(row=1, column=i).value:
            for j in range(2, ws.max_row + 1):
                if ws.cell(row=j, column=i).value is not None:
                    devices.append(str(ws.cell(row=j, column=i).value))
    return devices


def get_commands_by_sheet(datafile, sheet):
    wb, ws = load_xl(datafile, sheet)
    commands = []
    for i in range(2, ws.max_row + 1):
        temp_list = []
        cmd = ws.cell(row=i, column=1).value
        par = ws.cell(row=i, column=2).value
        temp_list = [str(cmd), str(par)]
        commands.append(temp_list)
    return commands




# def get_environment(env):
        # BuiltIn().log_to_console("22222")
